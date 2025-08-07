import openpyxl
from datetime import datetime
from openpyxl.utils.datetime import from_excel

# 勤務表C（透析室）から予定を抽出する
def parse_schedule_from_sheet_c(filepath, selected_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]  # ✅ 一番左のシートを使用

    # 姓だけ取り出す（例: "大江 直義" → "大江"）
    surname = selected_name.split()[0] if " " in selected_name else selected_name[:2]
    schedule_entries = []

    # セルから日付を取得
    def get_date_from_cell(cell):
        value = cell.value
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, (int, float)):
            try:
                return from_excel(value).date()
            except Exception as e:
                print(f"[ERROR] Excel serial date変換失敗: {value} → {e}")
        elif isinstance(value, str):
            try:
                dt = datetime.strptime(value.strip(), "%m/%d")
                return dt.replace(year=2025).date()
            except ValueError:
                pass
        return None

    # イベント作成
    def create_event(date):
        return {
            "summary": "透析室早出",
            "start": {
                "dateTime": date.strftime("%Y-%m-%dT07:30:00"),
                "timeZone": "Asia/Tokyo"
            },
            "end": {
                "dateTime": date.strftime("%Y-%m-%dT16:15:00"),
                "timeZone": "Asia/Tokyo"
            },
            "description": f"origin=C\nstaff={selected_name}"
        }

    # 氏名は列3、日付は列2（3〜40行をスキャン）
    for row in range(3, 40):
        name_cell = ws.cell(row=row, column=3)
        if name_cell.value and surname in str(name_cell.value):
            date_cell = ws.cell(row=row, column=2)
            date = get_date_from_cell(date_cell)

            if date is None:
                print(f"[WARNING] 日付取得失敗 at row {row}, value={date_cell.value}")
                continue

            print(f"[DEBUG] Parsed date {date} from cell {date_cell.value} for {selected_name} in row {row}")
            schedule_entries.append(create_event(date))

    print(f"[DEBUG] Parsed {len(schedule_entries)} events from file C for {selected_name}")
    return schedule_entries


# 勤務表Cから職員名を抽出する
def extract_names_from_sheet_c(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    names = set()

    # 氏名が記載されている列3、行は3〜40を対象
    for row in range(3, 40):
        value = ws.cell(row=row, column=3).value
        if isinstance(value, str):
            cleaned_name = value.strip()
            if cleaned_name:
                names.add(cleaned_name)

    return sorted(names)

def get_schedule_month_from_sheet_c(filepath):
    import openpyxl
    from datetime import datetime
    from openpyxl.utils.datetime import from_excel
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    for row in range(3, 40):
        value = ws.cell(row=row, column=2).value
        if isinstance(value, datetime):
            return value.year, value.month
        elif isinstance(value, (int, float)):
            try:
                dt = from_excel(value)
                return dt.year, dt.month
            except:
                continue
        elif isinstance(value, str):
            try:
                dt = datetime.strptime(value.strip(), "%m/%d")
                return 2025, dt.month
            except:
                continue
    return None, None
