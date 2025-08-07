import openpyxl
from datetime import datetime, timedelta

# =============================
# 勤務表Bから勤務予定を抽出する
# =============================
def parse_schedule_from_sheet_b(filepath, selected_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]  # 一番左のシート

    surname = selected_name.split()[0] if " " in selected_name else selected_name[:2]
    schedule_entries = []

    def get_date_from_row(row_idx):
        cell_value = ws.cell(row=row_idx, column=1).value
        if isinstance(cell_value, datetime):
            return cell_value.date()
        elif isinstance(cell_value, str):
            try:
                dt = datetime.strptime(cell_value.strip(), "%m/%d")
                return dt.replace(year=2025).date()
            except ValueError:
                pass
        return None

    # 色コード（ARGB）
    YELLOW = "FFFFFF00"
    ORANGE = "FFFFC000"

    def create_event(date, summary, start_time_str, end_time_str):
        start_dt = datetime.combine(date, datetime.strptime(start_time_str, "%H:%M").time())
        end_dt = datetime.combine(date, datetime.strptime(end_time_str, "%H:%M").time())
        return {
            "summary": summary,
            "start": {
                "dateTime": start_dt.strftime("%Y-%m-%dT%H:%M:%S"),
                "timeZone": "Asia/Tokyo"
            },
            "end": {
                "dateTime": end_dt.strftime("%Y-%m-%dT%H:%M:%S"),
                "timeZone": "Asia/Tokyo"
            },
            "description": f"origin=B\nstaff={selected_name}"  # ← originタグ付き
        }

    # === エリアA：Q9:AA28（col 17~27, row 9~28）===
    for row in range(9, 29):
        for col in range(17, 28):
            cell = ws.cell(row=row, column=col)
            if cell.value and surname in str(cell.value):
                fill = cell.fill
                color = fill.fgColor.rgb if fill and fill.fgColor.type == "rgb" else None
                date = get_date_from_row(row)

                if date is None:
                    continue

                if color == YELLOW:
                    schedule_entries.append(create_event(date, "心外早出", "07:30", "17:15"))
                elif color == ORANGE:
                    schedule_entries.append(create_event(date, "心外早出", "08:00", "17:15"))

    # === エリアB：AQ9:AS28（col 43~45, row 9~28）===
    for row in range(9, 29):
        for col in range(43, 46):
            cell = ws.cell(row=row, column=col)
            if cell.value and surname in str(cell.value):
                fill = cell.fill
                color = fill.fgColor.rgb if fill and fill.fgColor.type == "rgb" else None
                date = get_date_from_row(row)

                if date is None:
                    continue

                if color == ORANGE:
                    schedule_entries.append(create_event(date, "hinotori早出", "08:00", "17:15"))

    print(f"[DEBUG] Parsed {len(schedule_entries)} events from file B for {selected_name}")
    print(f"[DEBUG] 勤務表B抽出結果: {schedule_entries}")
    return schedule_entries

# ==========================================
# 勤務表Bから職員名候補を抽出（Aがないとき用）
# ==========================================
def extract_names_from_sheet_b(filepath):

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    names = set()

    # エリアA: Q9:AA28 → col=17~27, row=9~28
    for row in range(9, 29):
        for col in range(17, 28):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str):
                names.add(value.strip())

    # エリアB: AQ9:AS28 → col=43~45, row=9~28
    for row in range(9, 29):
        for col in range(43, 46):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str):
                names.add(value.strip())

    print(f"[DEBUG] 勤務表Bから抽出された職員名: {sorted(names)}")
    return sorted(names)

def get_schedule_month_from_sheet_b(filepath):
    import openpyxl
    from datetime import datetime
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    for row in range(9, 29):
        value = ws.cell(row=row, column=1).value
        if isinstance(value, datetime):
            return value.year, value.month
        elif isinstance(value, str):
            try:
                dt = datetime.strptime(value.strip(), "%m/%d")
                return 2025, dt.month  # 必要なら年は自動判定
            except:
                continue
    return None, None
